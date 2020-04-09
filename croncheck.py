from xml_parser import XML, XMLParser
import openpyxl

root_url = "/var/www/Ozone/Ozone/static/xml/"
filenames = [
    "mbt",
    "kbt"
]

format = ".xml"


for filename in filenames:
    parser = XMLParser(root_url + filename + format)

    offers = parser.ozon_offers_list()
    titles = parser.get_all_keys()
    categories = parser.get_all_categories()

    titles.sort(key = lambda x: str(offers[0:100]).count(x), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("offers")

    for key in titles:
        ws.cell(1, titles.index(key)+2, key)

    row = 2
    for offer in offers:
        ws.cell(row, 1, str(categories[int(offer["categoryId"])]))
        col = 2
        for key in titles:
            try:
                ws.cell(row, col, str(offer[key]))
                col += 1
            except:
                col += 1
        row += 1

    wb.save(root_url + filename + ".xlsx")
